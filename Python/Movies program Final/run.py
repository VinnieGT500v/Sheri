from openpyxl import load_workbook
class start:
	def __init__(self):
		self.movie_selected=""
		self.movie_list={}
		self.movie_seats=["A1","A2","A3","A4","A5","B1","B2","B3","B4","B5","C1","C2","C3","C4","C5","D1","D2","D3","D4","D5","E1","E2","E3","E4","E5"]
		self.open_cinema()

	def book_seat(self):
		current_movie = load_workbook(filename = "Cinemas.xlsx")
		current_sheet = current_movie["{}".format(self.movie_selected)]
		seat_list = []
		count = 1
		print("\t\t\t\t back of cinema")
		print("\t\t\tA\tB\tC\tD\tE")
		for seats in current_sheet.iter_rows(min_row=1,min_col=1,max_col=5,values_only=True):
		 	print()
		 	print("\t{}\t\t{}\t{}\t{}\t{}\t{}".format(count,seats[0],seats[1],seats[2],seats[3],seats[4]))
		 	count += 1
		while(True):
			choice = input("seat: ")
			if choice == "done":
				break
			else: 
				if choice not in self.movie_seats:
					print("invalid seat number")
				else:
					if choice in seat_list:
						print("seat already booked")
					else:
						if current_sheet[choice].value == "X":
							print("seat {} is already booked".format(choice))
						else:
							print("{} booked".format(choice))
							seat_list.append(choice)
		for x in seat_list:
			current_sheet[x] = "X"

		current_movie.save(filename = "Cinemas.xlsx")
		self.interface()


	def select_movie(self):
		print("movie options")
		for key,value in self.movie_list.items():
			print("\t{}. {}".format(key,value))
		while(True):
			choice=int(input("option: "))
			if choice <1 or choice > len(self.movie_list):
				print("invalid selection")
			else:
				self.movie_selected = self.movie_list["{}".format(choice)]
				self.sheet = self.workbook["{}".format(self.movie_selected)]
				break
		self.interface()
		
	def interface(self):
		print("movie_selected: {}".format(self.movie_selected))
		print("Menue options")
		if self.movie_selected!="":
			print("\t1. select a different movie")
		else:
			print("\t1. select a movie")
		print("\t2. book a seat")
		print("\t3. exit")
		choice=int(input("Option: "))
		if choice < 1 or choice > 3:
			print("invalid selection\n")
			self.interface()
		if choice == 1:
			self.select_movie()
		elif choice == 2:
			self.book_seat()
		elif choice == 3:
			exit()

	def open_cinema(self):
		self.workbook=load_workbook(filename="Cinemas.xlsx")
		count=1
		for x in self.workbook.sheetnames:
			self.movie_list["{}".format(count)]=x
			count+=1
		self.workbook.save(filename="Cinemas.xlsx")
		

run=start()
run.interface()
