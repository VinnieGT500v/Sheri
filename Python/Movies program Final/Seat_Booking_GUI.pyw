import tkinter as tk
from openpyxl import load_workbook
class GUI:
	def __init__(self):
		self.window = tk.Tk()
		self.window.geometry("500x200")
		self.window.grid_columnconfigure(0,weight = 1)
		self.window.grid_columnconfigure(1,weight = 1)
		self.window.grid_columnconfigure(2,weight = 1)
		self.window.grid_columnconfigure(3,weight = 1)
		self.window.grid_columnconfigure(4,weight = 1)

	def build_interface(self):
		self.movie_selected = tk.Label(master = self.window, font = (None, 20, "bold"))
		self.movie_selected.grid(row = 0, column = 1, columnspan = 3, sticky = "news")
		select_movie_btn = tk.Button(master = self.window, text = "Select Movie", command = self.select_movie)
		select_movie_btn.grid(row = 1, column = 0, columnspan = 2, sticky = "e")
		book_seat_btn = tk.Button(master = self.window, text = "Book Seat")
		book_seat_btn.grid(row = 1, column = 3, columnspan = 2, sticky = "w")
		self.window.mainloop()
	def select_movie(self):
		self.movie_selected["text"] = "hello world"

	def get_movie_list(self):
		workbook=load_workbook(filename="Cinemas.xlsx")
		for x in workbook.sheetnames:
			pass
		
		workbook.save(filename="Cinemas.xlsx")
run = GUI()
run.build_interface()